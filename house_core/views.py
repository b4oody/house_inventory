from decimal import Decimal

import openpyxl
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, HttpRequest
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views import generic
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Side, Border
from openpyxl.utils import get_column_letter

from house_core.forms import UserRegistrationForm, CreateRoomForm, CreateItemForm, ItemFilterForm
from house_core.models import Item, Apartment, Room, User, Category, Tag


def pagination(request, model_of_list):
    """
    :param request:
    :param model_of_list:
    :return:
    """
    paginator = Paginator(model_of_list, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return page_obj


def index_page_view(request: HttpRequest) -> HttpResponse:
    return render(request, "index/index.html")


class ProfilePageView(generic.DetailView):
    model = User
    template_name = "profile/profile.html"


@login_required
def report_page_view(request: HttpRequest) -> HttpResponse:
    apartments = Apartment.objects.filter(user=request.user)
    rooms = Room.objects.filter(apartment__in=apartments)
    items = Item.objects.filter(room__in=rooms)

    total_value = items.aggregate(
        total_items=Count("item_name"),
        total_purchase_price=Sum("purchase_price"),
        total_current_price=Sum("current_price"),
    )
    total_purchase_price = total_value.get("total_purchase_price") or Decimal("0")
    total_current_price = total_value.get("total_current_price") or Decimal("0")

    if total_purchase_price == Decimal("0"):
        difference_total = Decimal("0")
    else:
        difference_total = ((total_purchase_price - total_current_price) /
                            total_purchase_price) * 100

    apartments_sums = Apartment.objects.filter(user=request.user).annotate(
        total_purchase_price=Sum("rooms__items__purchase_price"),
        total_current_price=Sum("rooms__items__current_price"),
        total_items=Count("rooms__items"),
    )

    rooms_sums = Room.objects.filter(apartment__user=request.user).annotate(
        total_purchase_price=Sum("items__purchase_price"),
        total_current_price=Sum("items__current_price"),
        total_items=Count("items"),
    )

    context = {
        "total_items": total_value.get("total_items", 0),
        "total_current_price": total_current_price,
        "total_purchase_price": total_purchase_price,
        "difference_total": difference_total,
        "apartments_sums": apartments_sums
    }

    return render(
        request,
        "reports/reports.html",
        context=context,
    )


@login_required
def items_page_view(request: HttpRequest) -> HttpResponse:
    items = Item.objects.filter(room__apartment__user=request.user)
    exclude = {"id", "photo_url", "created_at"}

    fields = [
        field for field in Item._meta.concrete_fields
        if field.name not in exclude
    ]

    form = ItemFilterForm(request.GET, user=request.user)
    if form.is_valid():
        apartment = form.cleaned_data["apartment"]
        room = form.cleaned_data["room"]
        query = form.cleaned_data["query"]

        if apartment and apartment != "all":
            items = items.filter(room__apartment=apartment)

        if room and room != "all":
            items = items.filter(room=room)

        if query:
            items = items.filter(
                Q(item_name__icontains=query) | Q(brand__icontains=query)
            )

    page_obj = pagination(request, items)
    context = {
        "item_fields": fields,
        "page_obj": page_obj,
        "items": page_obj,
        "form": form,
    }

    return render(
        request,
        "items/items.html",
        context=context
    )


@login_required
def apartments_page_view(request: HttpRequest) -> HttpResponse:
    exclude = {"id", "user", "created_at"}
    apartments = Apartment.objects.filter(user=request.user)
    page_obj = pagination(request, apartments)

    fields = [
        field for field in Apartment._meta.concrete_fields
        if field.name not in exclude
    ]

    context = {
        "apartments_fields": fields,
        "page_obj": page_obj,
        "apartments": page_obj
    }
    return render(
        request,
        "apartments/apartments.html",
        context=context
    )


@login_required
def apartment_page_view(request: HttpRequest, pk: id) -> HttpResponse:
    apartment = Apartment.objects.get(pk=pk)

    exclude = {"id", "apartment", "created_at"}
    rooms = Room.objects.filter(apartment=apartment)
    page_obj = pagination(request, rooms)

    fields = [
        field for field in Room._meta.concrete_fields
        if field.name not in exclude
    ]

    context = {
        "apartment": apartment,
        "room_fields": fields,
        "page_obj": page_obj,
        "rooms": page_obj
    }
    return render(
        request,
        "apartments/apartment.html",
        context=context
    )


class ApartmentUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Apartment
    template_name = "apartments/update_apartment_form.html"
    fields = [
        "apartment_name",
        "address",
        "apartment_description",
        "purchase_price"
    ]

    def get_success_url(self):
        return reverse_lazy(
            "house_core:pk_apartment_view",
            kwargs={"pk": self.object.pk}
        )


class ApartmentDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Apartment
    template_name = "apartments/delete_apartment_form.html"
    success_url = reverse_lazy("house_core:apartments_view")


class UserRegistrationView(generic.CreateView):
    form_class = UserRegistrationForm
    template_name = "registration/register.html"
    success_url = reverse_lazy("house_core:apartments_view")

    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        return response


class ApartmentCreateView(LoginRequiredMixin, generic.CreateView):
    model = Apartment
    template_name = "apartments/create_apartment_form.html"
    fields = [
        "apartment_name",
        "address",
        "apartment_description",
        "purchase_price"
    ]

    def get_success_url(self):
        return reverse_lazy(
            "house_core:pk_apartment_view",
            kwargs={"pk": self.object.pk}
        )

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


@login_required
def room_page_view(request: HttpRequest, pk: id) -> HttpResponse:
    room = Room.objects.get(pk=pk)

    exclude = {"id", "photo_url", "room", "created_at"}
    items = Item.objects.filter(room=room)
    page_obj = pagination(request, items)

    fields = [
        field for field in Item._meta.concrete_fields
        if field.name not in exclude
    ]

    context = {
        "room": room,
        "item_fields": fields,
        "page_obj": page_obj,
        "items": page_obj
    }
    return render(
        request,
        "rooms/room.html",
        context=context
    )


class RoomUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Room
    template_name = "rooms/update_room_form.html"
    fields = [
        "room_name",
        "room_description",
        "area_m2",
    ]

    def get_success_url(self):
        return reverse_lazy(
            "house_core:pk_room_view",
            kwargs={"pk": self.object.pk}
        )


class RoomDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Room
    template_name = "rooms/delete_room_form.html"
    success_url = reverse_lazy("house_core:apartments_view")


class RoomCreateView(LoginRequiredMixin, generic.CreateView):
    model = Room
    form_class = CreateRoomForm
    template_name = "rooms/create_room_form.html"

    def get_success_url(self):
        return reverse_lazy(
            "house_core:pk_room_view",
            kwargs={"pk": self.object.pk}
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs


class ItemUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Item
    template_name = "items/update_item_form.html"
    fields = "__all__"

    success_url = reverse_lazy("house_core:items_view")


class ItemDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Item
    template_name = "items/delete_item_form.html"
    success_url = reverse_lazy("house_core:items_view")


class ItemCreateView(LoginRequiredMixin, generic.CreateView):
    model = Item
    form_class = CreateItemForm
    template_name = "items/create_item_form.html"
    success_url = reverse_lazy("house_core:items_view")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs


@login_required
def all_rooms_page_view(request: HttpRequest) -> HttpResponse:
    exclude = {"id", "created_at"}
    rooms = Room.objects.filter(apartment__user=request.user)
    page_obj = pagination(request, rooms)

    fields = [
        field for field in Room._meta.concrete_fields
        if field.name not in exclude
    ]

    context = {
        "rooms_fields": fields,
        "page_obj": page_obj,
        "rooms": page_obj
    }
    return render(
        request,
        "rooms/rooms.html",
        context=context
    )


@login_required
def export_to_excel(request):
    wb = openpyxl.Workbook()

    # Створюємо окремі листи для кожної моделі
    sheets_data = [
        ("Apartments", Apartment.objects.filter(user=request.user), [
            ("apartment_name", "Apartment Name"),
            ("address", "Address"),
            ("apartment_description", "Description"),
            ("purchase_price", "Purchase Price ₴"),
            ("created_at", "Created At")
        ]),
        ("Rooms", Room.objects.filter(apartment__user=request.user), [
            ("room_name", "Room Name"),
            ("room_description", "Description"),
            ("area_m2", "Area (m²)"),
            ("created_at", "Created At"),
            ("apartment.apartment_name", "Apartment Name")  # Доступ до зв'язку ForeignKey
        ]),
        ("Items", Item.objects.filter(room__apartment__user=request.user), [
            ("item_name", "Item Name"),
            ("brand", "Brand"),
            ("quantity", "Quantity"),
            ("purchase_price", "Purchase Price"),
            ("current_price", "Current Price"),
            ("purchase_date", "Purchase Date"),
            ("warranty_until", "Warranty Until"),
            ("room.room_name", "Room Name")  # Доступ до зв'язку ForeignKey
        ]),
        ("Categories", Category.objects.all(), [
            ("category_name", "Category Name")
        ]),
        ("Tags", Tag.objects.all(), [
            ("tag_name", "Tag Name")
        ])
    ]

    for sheet_name, model_data, fields in sheets_data:
        ws = wb.create_sheet(title=sheet_name)

        # Заголовки стовпців
        columns = [field[1] for field in fields]
        for col_num, column_title in enumerate(columns, 1):
            c = ws.cell(row=1, column=col_num)
            c.value = column_title
            c.alignment = Alignment(horizontal="center", vertical="center")  # Центруємо заголовок
            c.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

        # Додавання даних
        for row_num, instance in enumerate(model_data, 2):
            for col_num, (field, _) in enumerate(fields, 1):
                value = getattr(instance, field, None)

                # Якщо поле є зовнішнім ключем або має інші поля, витягуємо потрібне значення
                if hasattr(value, '__str__'):  # Перевірка, чи це об'єкт, що можна привести до рядка
                    value = str(value)

                ws.cell(row=row_num, column=col_num, value=value).border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin")
                )

        # Розтягування стовпців
        for col_num in range(1, len(columns) + 1):
            max_length = 0
            column = openpyxl.utils.get_column_letter(col_num)
            for row in range(1, len(model_data) + 2):
                cell_value = ws[f'{column}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Відповідь користувачу
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data_export.xlsx"'

    wb.save(response)
    return response
