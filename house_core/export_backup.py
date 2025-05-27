import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl.styles import Side, Border, Alignment

from house_core.models import Tag, Category, Item, Room, Apartment


@login_required
def export_to_excel(request):
    wb = openpyxl.Workbook()

    # Створюємо окремі листи для кожної моделі
    sheets_data = [
        ("Apartments", Apartment.objects.filter(user=request.user), [
            ("apartment_name", "Apartment Name"),
            ("address", "Address"),
            ("apartment_description", "Description"),
            ("purchase_price", "Purchase Price ₴"),
            ("created_at", "Created At")
        ]),
        ("Rooms", Room.objects.filter(apartment__user=request.user), [
            ("room_name", "Room Name"),
            ("room_description", "Description"),
            ("area_m2", "Area (m²)"),
            ("created_at", "Created At"),
            ("apartment.apartment_name", "Apartment Name")  # Доступ до зв'язку ForeignKey
        ]),
        ("Items", Item.objects.filter(room__apartment__user=request.user), [
            ("item_name", "Item Name"),
            ("brand", "Brand"),
            ("quantity", "Quantity"),
            ("purchase_price", "Purchase Price"),
            ("current_price", "Current Price"),
            ("purchase_date", "Purchase Date"),
            ("warranty_until", "Warranty Until"),
            ("room.room_name", "Room Name")  # Доступ до зв'язку ForeignKey
        ]),
        ("Categories", Category.objects.all(), [
            ("category_name", "Category Name")
        ]),
        ("Tags", Tag.objects.all(), [
            ("tag_name", "Tag Name")
        ])
    ]

    for sheet_name, model_data, fields in sheets_data:
        ws = wb.create_sheet(title=sheet_name)

        # Заголовки стовпців
        columns = [field[1] for field in fields]
        for col_num, column_title in enumerate(columns, 1):
            c = ws.cell(row=1, column=col_num)
            c.value = column_title
            c.alignment = Alignment(horizontal="center", vertical="center")  # Центруємо заголовок
            c.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

        # Додавання даних
        for row_num, instance in enumerate(model_data, 2):
            for col_num, (field, _) in enumerate(fields, 1):
                value = getattr(instance, field, None)

                # Якщо поле є зовнішнім ключем або має інші поля, витягуємо потрібне значення
                if hasattr(value, '__str__'):  # Перевірка, чи це об'єкт, що можна привести до рядка
                    value = str(value)

                ws.cell(row=row_num, column=col_num, value=value).border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin")
                )

        # Розтягування стовпців
        for col_num in range(1, len(columns) + 1):
            max_length = 0
            column = openpyxl.utils.get_column_letter(col_num)
            for row in range(1, len(model_data) + 2):
                cell_value = ws[f'{column}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Відповідь користувачу
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="data_export.xlsx"'

    wb.save(response)
    return response
